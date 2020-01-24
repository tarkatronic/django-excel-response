from __future__ import absolute_import, unicode_literals

from collections import OrderedDict

import openpyxl
import six
from django.test import TestCase
from openpyxl.styles import Font

from excel_response import response

from .models import TestModel


class ExcelResponseCSVTest(TestCase):

    def test_force_csv(self):
        r = response.ExcelResponse([['a']], force_csv=True)
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')

    def test_force_csv_dict(self):
        r = response.ExcelResponse({'worksheet': [['a']]}, force_csv=True)
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')

    def test_exceeding_row_limit_with_list_creates_csv(self):
        old_limit = response.ROW_LIMIT
        response.ROW_LIMIT = 2
        r = response.ExcelResponse(
            [['a'], ['b'], ['c']]
        )
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')
        response.ROW_LIMIT = old_limit

    def test_exceeding_row_limit_with_dict_creates_csv(self):
        old_limit = response.ROW_LIMIT
        response.ROW_LIMIT = 2
        r = response.ExcelResponse(
            {'worksheet': [['a'], ['b'], ['c']]}
        )
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')
        response.ROW_LIMIT = old_limit

    def test_exceeding_column_limit_with_list_creates_csv(self):
        old_limit = response.COL_LIMIT
        response.COL_LIMIT = 2
        r = response.ExcelResponse(
            [['a', 'b', 'c']]
        )
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')
        response.COL_LIMIT = old_limit

    def test_exceeding_column_limit_with_dict_creates_csv(self):
        old_limit = response.COL_LIMIT
        response.COL_LIMIT = 2
        r = response.ExcelResponse({
            'worksheet_1': [['a']],
            'worksheet_2': [['a', 'b', 'c']]
            })
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')
        response.COL_LIMIT = old_limit

    def test_exceeding_row_limit_with_model_creates_csv(self):
        old_limit = response.ROW_LIMIT
        response.ROW_LIMIT = 2
        TestModel.objects.create(text='a', number='1')
        TestModel.objects.create(text='b', number='2')
        TestModel.objects.create(text='c', number='3')
        r = response.ExcelResponse(
            TestModel.objects.all()
        )
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')
        response.ROW_LIMIT = old_limit

    def test_exceeding_row_limit_with_dict_of_model_creates_csv(self):
        old_limit = response.ROW_LIMIT
        response.ROW_LIMIT = 2
        TestModel.objects.create(text='a', number='1')
        TestModel.objects.create(text='b', number='2')
        TestModel.objects.create(text='c', number='3')
        r = response.ExcelResponse(
            {'worksheet': TestModel.objects.all()}
        )
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')
        response.ROW_LIMIT = old_limit

    def test_exceeding_column_limit_with_model_creates_csv(self):
        old_limit = response.COL_LIMIT
        response.COL_LIMIT = 2
        TestModel.objects.create(text='a', number='1')
        r = response.ExcelResponse(
            TestModel.objects.all()
        )
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')
        response.COL_LIMIT = old_limit

    def test_exceeding_column_limit_with_dict_of_model_creates_csv(self):
        old_limit = response.COL_LIMIT
        response.COL_LIMIT = 2
        TestModel.objects.create(text='a', number='1')
        r = response.ExcelResponse(
            {'worksheet': TestModel.objects.all()}
        )
        # Call this so that all of the data gets resolved
        r.content
        self.assertEqual(r['content-type'], 'text/csv; charset=utf8')
        response.COL_LIMIT = old_limit

    def test_csv_from_list(self):
        r = response.ExcelResponse(
            [['a', 'b', 'c'], [1, 2, 3], [4, 5, 6]],
            force_csv=True
        )
        output = r.getvalue()
        self.assertEqual(
            output,
            b'a,b,c\r\n1,2,3\r\n4,5,6\r\n'
        )

    def test_csv_from_list_of_dicts(self):
        r = response.ExcelResponse(
            [
                OrderedDict([('a', 1), ('b', 2), ('c', 3)]),  # OrderedDict ensures the order of our headers & output
                {'a': 4, 'b': 5, 'c': 6}
            ],
            force_csv=True
        )
        output = r.getvalue()
        self.assertEqual(
            output,
            b'a,b,c\r\n1,2,3\r\n4,5,6\r\n'
        )

    def test_csv_from_dict(self):
        r = response.ExcelResponse(
            OrderedDict([
                ('worksheet_1', [['a', 'b', 'c'], [1, 2, 3], [4, 5, 6]]),
                ('worksheet_2', [['e', 'f'], [7, 8]])
            ]),
            force_csv=True
        )
        output = r.getvalue()
        self.assertEqual(
            output,
            b'\r\n\r\nworksheet_1\r\n\r\na,b,c\r\n1,2,3\r\n4,5,6\r\n\r\n\r\nworksheet_2\r\n\r\ne,f\r\n7,8\r\n'
        )

    def test_queryset_values_limits_output_columns(self):
        TestModel.objects.create(text='a', number='1')
        TestModel.objects.create(text='b', number='2')
        TestModel.objects.create(text='c', number='3')
        r = response.ExcelResponse(
            TestModel.objects.all().values('text'),
            force_csv=True
        )
        output = r.getvalue()
        self.assertEqual(
            output,
            b'text\r\na\r\nb\r\nc\r\n'
        )

    def test_csv_output_with_empty_list(self):
        r = response.ExcelResponse([], force_csv=True)
        output = r.getvalue()
        self.assertEqual(output, b'')

    def test_csv_output_with_empty_dict(self):
        r = response.ExcelResponse({}, force_csv=True)
        output = r.getvalue()
        self.assertEqual(output, b'')

    def test_csv_output_with_empty_queryset(self):
        r = response.ExcelResponse(TestModel.objects.none(), force_csv=True)
        output = r.getvalue()
        self.assertEqual(output, b'')


class ExcelResponseExcelTest(TestCase):

    def test_create_excel_from_list(self):
        r = response.ExcelResponse(
            [['a', 'b', 'c'], [1, 2, 3], [4, 5, 6]]
        )
        output = six.BytesIO(r.getvalue())
        # This should theoretically raise errors if it's not a valid spreadsheet
        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb.active
        self.assertEqual((ws['A1'].value, ws['B1'].value, ws['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((ws['A2'].value, ws['B2'].value, ws['C2'].value), (1, 2, 3))

    def test_create_excel_from_list_of_dicts(self):
        r = response.ExcelResponse(
            [
                OrderedDict([('a', 1), ('b', 2), ('c', 3)]),  # OrderedDict ensures the order of our headers & output
                {'a': 4, 'b': 5, 'c': 6}
            ]
        )
        output = six.BytesIO(r.getvalue())
        openpyxl.load_workbook(output, read_only=True)

    def test_create_excel_from_queryset(self):
        TestModel.objects.create(text='a', number='1')
        TestModel.objects.create(text='b', number='2')
        TestModel.objects.create(text='c', number='3')
        r = response.ExcelResponse(
            TestModel.objects.all()
        )
        output = six.BytesIO(r.getvalue())
        openpyxl.load_workbook(output, read_only=True)

    def test_header_font_is_applied(self):
        f = Font(name='Windings')
        r = response.ExcelResponse(
            [['a', 'b', 'c'], [1, 2, 3], [4, 5, 6]],
            header_font=f
        )
        output = six.BytesIO(r.getvalue())
        book = openpyxl.load_workbook(output, read_only=True)
        sheet = book.active
        cell = sheet['A1']
        self.assertEqual(cell.font.name, 'Windings')

    def test_data_font_is_applied(self):
        f = Font(name='Windings')
        r = response.ExcelResponse(
            [['a', 'b', 'c'], [1, 2, 3], [4, 5, 6]],
            data_font=f
        )
        output = six.BytesIO(r.getvalue())
        book = openpyxl.load_workbook(output, read_only=True)
        sheet = book.active
        cell = sheet['A2']
        self.assertEqual(cell.font.name, 'Windings')

    def test_create_excel_with_guess_types_on(self):
        r = response.ExcelResponse(
            [['a', 'b', 'c'], ['2018032710050111540290000000000720000000023', 2, 3], [4, 5, 6]]
        )
        output = six.BytesIO(r.getvalue())
        # This should theoretically raise errors if it's not a valid spreadsheet
        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb.active
        self.assertEqual((ws['A1'].value, ws['B1'].value, ws['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((ws['A2'].value, ws['B2'].value, ws['C2'].value), (2.018032710050112e+42, 2, 3))

    def test_create_excel_with_guess_types_off(self):
        r = response.ExcelResponse(
            [['a', 'b', 'c'], ['2018032710050111540290000000000720000000023', 2, 3], [4, 5, 6]], guess_types=False
        )
        output = six.BytesIO(r.getvalue())
        # This should theoretically raise errors if it's not a valid spreadsheet
        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb.active
        self.assertEqual((ws['A1'].value, ws['B1'].value, ws['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((ws['A2'].value, ws['B2'].value, ws['C2'].value),
                         ('2018032710050111540290000000000720000000023', 2, 3))

    def test_output_with_empty_list(self):
        r = response.ExcelResponse([])
        output = r.getvalue()
        self.assertEqual(output, b'')

    def test_output_with_empty_queryset(self):
        r = response.ExcelResponse(TestModel.objects.none())
        output = r.getvalue()
        self.assertEqual(output, b'')


class ExcelResponseMultiSheetExcelTest(TestCase):
    def setUp(self):
        self.one_worksheet_dict = {'worksheet': [['a', 'b', 'c'], [1, 2, 3], [4, 5, 6]]}
        self.multi_worksheet_dict = OrderedDict([
            ('worksheet_1', [['a', 'b', 'c'], [1, 2, 3], [4, 5, 6]]),
            ('worksheet_2', [['d', 'e'], [7, 8]]),
        ])

    def test_create_excel_from_dict_one_key(self):
        r = response.ExcelResponse(self.one_worksheet_dict)
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb.active
        self.assertEqual(ws.title, 'worksheet')
        self.assertEqual((ws['A1'].value, ws['B1'].value, ws['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((ws['A2'].value, ws['B2'].value, ws['C2'].value), (1, 2, 3))

    def test_create_excel_from_dict_one_key_ignore_worksheet_name(self):
        r = response.ExcelResponse(self.one_worksheet_dict, worksheet_name='ignored')
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb.active
        self.assertEqual(ws.title, 'worksheet')
        self.assertEqual((ws['A1'].value, ws['B1'].value, ws['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((ws['A2'].value, ws['B2'].value, ws['C2'].value), (1, 2, 3))

    def test_create_excel_from_dict_multi_key(self):
        r = response.ExcelResponse(self.multi_worksheet_dict)
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output, read_only=True)

        sheet_1 = wb.worksheets[0]
        self.assertEqual(sheet_1.title, 'worksheet_1')
        self.assertEqual((sheet_1['A1'].value, sheet_1['B1'].value, sheet_1['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((sheet_1['A2'].value, sheet_1['B2'].value, sheet_1['C2'].value), (1, 2, 3))

        sheet_2 = wb.worksheets[1]
        self.assertEqual(sheet_2.title, 'worksheet_2')
        self.assertEqual((sheet_2['A1'].value, sheet_2['B1'].value), ('d', 'e'))
        self.assertEqual((sheet_2['A2'].value, sheet_2['B2'].value), (7, 8))

    def test_create_excel_from_list_of_dicts(self):
        r = response.ExcelResponse(OrderedDict([
            ('worksheet_1', [
                OrderedDict([('a', 1), ('b', 2), ('c', 3)]),  # OrderedDict ensures the order of our headers & output
                {'a': 4, 'b': 5, 'c': 6}
            ]),
            ('worksheet_2', [OrderedDict([('d', 7), ('e', 8)])])
        ]))
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output, read_only=True)

        sheet_1 = wb.worksheets[0]
        self.assertEqual(sheet_1.title, 'worksheet_1')
        self.assertEqual((sheet_1['A1'].value, sheet_1['B1'].value, sheet_1['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((sheet_1['A2'].value, sheet_1['B2'].value, sheet_1['C2'].value), (1, 2, 3))
        self.assertEqual((sheet_1['A3'].value, sheet_1['B3'].value, sheet_1['C3'].value), (4, 5, 6))

        sheet_2 = wb.worksheets[1]
        self.assertEqual(sheet_2.title, 'worksheet_2')
        self.assertEqual((sheet_2['A1'].value, sheet_2['B1'].value), ('d', 'e'))
        self.assertEqual((sheet_2['A2'].value, sheet_2['B2'].value), (7, 8))

    def test_create_excel_from_dict_of_querysets(self):
        TestModel.objects.create(text='a', number='4')
        TestModel.objects.create(text='b', number='5')
        TestModel.objects.create(text='c', number='6')
        r = response.ExcelResponse(
            {'worksheet': TestModel.objects.all()}
        )
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output, read_only=True)
        self.assertEqual(len(wb.worksheets), 1)

        worksheet = wb.worksheets[0]
        first_row = (worksheet['A1'].value, worksheet['B1'].value, worksheet['C1'].value, worksheet['D1'].value)
        self.assertIn('id', first_row)
        self.assertIn('text', first_row)
        self.assertIn('number', first_row)
        self.assertIn('timestamp', first_row)

        second_row = (worksheet['A2'].value, worksheet['B2'].value, worksheet['C2'].value, worksheet['D2'].value)
        self.assertIn(1, second_row)
        self.assertIn(4, second_row)
        self.assertIn('a', second_row)

    def test_header_font_is_applied(self):
        f = Font(name='Windings')
        r = response.ExcelResponse(
            self.one_worksheet_dict,
            header_font=f
        )
        output = six.BytesIO(r.getvalue())
        book = openpyxl.load_workbook(output, read_only=True)
        sheet = book.active
        cell = sheet['A1']
        self.assertEqual(cell.font.name, 'Windings')

    def test_data_font_is_applied(self):
        f = Font(name='Windings')
        r = response.ExcelResponse(
            self.one_worksheet_dict,
            data_font=f
        )
        output = six.BytesIO(r.getvalue())
        book = openpyxl.load_workbook(output, read_only=True)
        sheet = book.active
        cell = sheet['A2']
        self.assertEqual(cell.font.name, 'Windings')

    def test_create_excel_with_guess_types_on(self):
        r = response.ExcelResponse(
            {'worksheet': [['a', 'b', 'c'], ['2018032710050111540290000000000720000000023', 2, 3], [4, 5, 6]]}
        )
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb.active
        self.assertEqual((ws['A1'].value, ws['B1'].value, ws['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((ws['A2'].value, ws['B2'].value, ws['C2'].value), (2.018032710050112e+42, 2, 3))

    def test_create_excel_with_guess_types_off(self):
        r = response.ExcelResponse(
            {'worksheet': [['a', 'b', 'c'], ['2018032710050111540290000000000720000000023', 2, 3], [4, 5, 6]]},
            guess_types=False
        )
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output, read_only=True)
        ws = wb.active
        self.assertEqual((ws['A1'].value, ws['B1'].value, ws['C1'].value), ('a', 'b', 'c'))
        self.assertEqual((ws['A2'].value, ws['B2'].value, ws['C2'].value),
                         ('2018032710050111540290000000000720000000023', 2, 3))

    def test_output_with_empty_dict(self):
        r = response.ExcelResponse({})
        output = r.getvalue()
        self.assertEqual(output, b'')

    def test_output_with_empty_value_in_dict(self):
        r = response.ExcelResponse({'name': []})
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        self.assertIn('name', wb.sheetnames)
        self.assertEqual(len(list(ws.columns)), 1)
        self.assertEqual(len(list(ws.rows)), 1)
        self.assertFalse(ws['A1'].value)

    def test_output_with_empty_qs_in_dict(self):
        r = response.ExcelResponse({'name': TestModel.objects.none()})
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        self.assertIn('name', wb.sheetnames)
        self.assertEqual(len(list(ws.columns)), 1)
        self.assertEqual(len(list(ws.rows)), 1)
        self.assertFalse(ws['A1'].value)

    def test_create_excel_with_not_str_keys(self):
        r = response.ExcelResponse(
            {
                1: [['a', 'b'], [1, 2]],
                ('c', 'd'): [['f']]
             },
            guess_types=False
        )
        output = six.BytesIO(r.getvalue())
        wb = openpyxl.load_workbook(output, read_only=True)
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn('1', wb.sheetnames)


class CBVTest(TestCase):

    def test_excel_view(self):
        TestModel.objects.create(text='a', number='1')
        TestModel.objects.create(text='b', number='2')
        TestModel.objects.create(text='c', number='3')
        response = self.client.get('/test/')
        output = six.BytesIO(response.content)
        openpyxl.load_workbook(output, read_only=True)
